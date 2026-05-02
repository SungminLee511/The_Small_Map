"""Importer: Seoul (Mapo-gu) public toilets.

Source: 전국공중화장실표준데이터 (National Public Toilet Standard Data)
Portal: https://www.data.go.kr/data/15012892/standard.do

**Schema verified Phase 1.2** (snapshot 2024-02-18, 6 089 rows nationwide):

- The standard does **not** publish a stable ``화장실관리번호`` column.
  We synthesise a deterministic external id from
  ``(road address | lot address | name)`` — see ``_row_external_id``.
- There is no ``남녀공용화장실여부`` flag. Gender is inferred from the
  per-fixture stall counts (``남성용-대변기수`` etc.).
- There is no ``장애인용화장실설치여부`` flag. Accessibility is true iff
  the sum of accessible-stall counts is positive.
- The Seoul snapshot ships **47 % of nationwide rows with WGS84 coords**.
  In Mapo-gu specifically **0/194 rows have coordinates** (subway-station
  toilets all have blank lat/lng) — they must be geocoded by address.
  We accept an injectable ``geocoder`` callable like
  ``seoul_smoking_areas`` and run it only when coords are missing.
- The download is delivered as XLSX as well as CP949 CSV. Both are
  supported.
"""

from __future__ import annotations

import csv
import hashlib
import io
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Awaitable, Callable, Sequence

import httpx

from app.importers.base import BaseImporter, ImportReport, POIInput
from app.models.poi import POIType

logger = logging.getLogger(__name__)


# --- Source columns (Korean). Verified against the 2024-02-18 snapshot. ---
COL_NAME = "화장실명"
COL_ROAD_ADDR = "소재지도로명주소"
COL_LOT_ADDR = "소재지지번주소"
COL_LAT = "WGS84위도"
COL_LNG = "WGS84경도"

# Per-fixture stall counts (used to infer gender + accessibility).
COL_MEN_STALLS = "남성용-대변기수"
COL_MEN_URINALS = "남성용-소변기수"
COL_MEN_ACC_STALLS = "남성용-장애인용대변기수"
COL_MEN_ACC_URINALS = "남성용-장애인용소변기수"
COL_WOMEN_STALLS = "여성용-대변기수"
COL_WOMEN_ACC_STALLS = "여성용-장애인용대변기수"

COL_HOURS = "개방시간"
COL_HOURS_DETAIL = "개방시간상세"
COL_BABY = "기저귀교환대유무"
COL_AS_OF = "데이터기준일자"

DISTRICT_KEYWORD = "마포구"

GeocoderFn = Callable[[str], Awaitable[tuple[float, float] | None]]


def _truthy(val) -> bool | None:
    """Korean public CSVs use 'Y'/'N'/'있음'/'없음'. Map to bool, None if blank."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s in {"", "-", "N/A", "NULL"}:
        return None
    if s in {"Y", "YES", "TRUE", "1", "있음", "유"}:
        return True
    if s in {"N", "NO", "FALSE", "0", "없음", "무"}:
        return False
    return None


def _to_int(val) -> int | None:
    if val is None or str(val).strip() == "":
        return None
    try:
        return int(float(str(val).strip()))
    except (TypeError, ValueError):
        return None


def _parse_as_of(val) -> datetime | None:
    if not val:
        return None
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _row_external_id(row: dict) -> str:
    """Deterministic id from ``(road | lot | name)``.

    The standard data has no stable PK column — sequential ``번호`` is
    not safe across snapshots. The address+name tuple is stable enough
    for our upsert semantics.
    """
    addr = (row.get(COL_ROAD_ADDR) or row.get(COL_LOT_ADDR) or "").strip()
    name = (row.get(COL_NAME) or "").strip()
    blob = f"{addr}|{name}".encode("utf-8")
    return "seoul-tol-" + hashlib.sha1(blob).hexdigest()[:16]


def _gender_from_counts(row: dict) -> str | None:
    men = (_to_int(row.get(COL_MEN_STALLS)) or 0) + (_to_int(row.get(COL_MEN_URINALS)) or 0)
    women = _to_int(row.get(COL_WOMEN_STALLS)) or 0
    if men > 0 and women > 0:
        return "separate"
    if men > 0:
        return "male_only"
    if women > 0:
        return "female_only"
    return None


def _accessibility_from_counts(row: dict) -> bool | None:
    cols = (COL_MEN_ACC_STALLS, COL_MEN_ACC_URINALS, COL_WOMEN_ACC_STALLS)
    seen_any = False
    total = 0
    for c in cols:
        n = _to_int(row.get(c))
        if n is not None:
            seen_any = True
            total += n
    if not seen_any:
        return None
    return total > 0


def _opening_hours(row: dict) -> str | None:
    """Combine ``개방시간`` + ``개방시간상세`` when both present."""
    coarse = (row.get(COL_HOURS) or "").strip()
    detail = (row.get(COL_HOURS_DETAIL) or "").strip()
    if detail:
        return detail
    if coarse:
        return coarse
    return None


class SeoulPublicToiletsImporter(BaseImporter):
    source_id = "seoul.public_toilets"
    poi_type = POIType.toilet
    cycle_days = 30

    def __init__(
        self,
        *,
        csv_path: str | Path | None = None,
        xlsx_path: str | Path | None = None,
        api_url: str | None = None,
        district_keyword: str = DISTRICT_KEYWORD,
        encoding: str = "cp949",
        geocoder: GeocoderFn | None = None,
    ):
        self.csv_path = Path(csv_path) if csv_path else None
        self.xlsx_path = Path(xlsx_path) if xlsx_path else None
        self.api_url = api_url
        self.district_keyword = district_keyword
        self.encoding = encoding
        self.geocoder = geocoder
        # Per-run cache so repeated identical addresses don't hammer Kakao.
        self._geocode_cache: dict[str, tuple[float, float] | None] = {}

    # --- raw fetch ---------------------------------------------------------

    async def fetch_raw(self) -> Sequence[dict]:
        if self.xlsx_path is not None:
            return self._read_xlsx(self.xlsx_path)
        if self.csv_path is not None:
            return self._read_csv(self.csv_path.read_bytes())
        if self.api_url is not None:
            async with httpx.AsyncClient(timeout=60.0) as client:
                resp = await client.get(self.api_url)
                resp.raise_for_status()
                payload = resp.json()
                if isinstance(payload, dict) and "data" in payload:
                    return payload["data"]
                if isinstance(payload, list):
                    return payload
                return []
        raise RuntimeError("Either csv_path / xlsx_path / api_url must be provided")

    def _read_csv(self, raw_bytes: bytes) -> list[dict]:
        # Try the configured encoding first, then fall back to UTF-8 (some
        # mirrors re-encode the file on the way out).
        for enc in (self.encoding, "utf-8-sig", "utf-8"):
            try:
                text = raw_bytes.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = raw_bytes.decode(self.encoding, errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return list(reader)

    def _read_xlsx(self, path: Path) -> list[dict]:
        # Lazy import: keeps openpyxl out of the hot path for users who only
        # ever feed CSVs.
        try:
            import openpyxl  # type: ignore
        except ImportError as e:  # pragma: no cover - exercised via run-time error
            raise RuntimeError(
                "openpyxl is required to read xlsx. Install with `pip install openpyxl`."
            ) from e

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            return []
        out: list[dict] = []
        for raw in rows_iter:
            row = {h: (str(v) if v is not None else "") for h, v in zip(header, raw)}
            out.append(row)
        return out

    # --- filters & geocoding ------------------------------------------------

    def _is_in_district(self, row: dict) -> bool:
        if not self.district_keyword:
            return True
        for col in (COL_ROAD_ADDR, COL_LOT_ADDR):
            v = row.get(col)
            if v and self.district_keyword in str(v):
                return True
        return False

    async def _geocode(self, address: str) -> tuple[float, float] | None:
        if not address:
            return None
        if address in self._geocode_cache:
            return self._geocode_cache[address]
        if self.geocoder is None:
            return None
        result = await self.geocoder(address)
        self._geocode_cache[address] = result
        return result

    # --- normalize ----------------------------------------------------------

    def normalize(self, raw: dict) -> POIInput | None:
        """Sync path used when a row already has WGS84 coordinates.

        Returns ``None`` when the row needs geocoding — in that case
        ``_normalize_async`` (called from our overridden ``run``) will
        re-process the row via the injected geocoder.
        """
        if not self._is_in_district(raw):
            return None

        try:
            lat = float(raw.get(COL_LAT))
            lng = float(raw.get(COL_LNG))
        except (TypeError, ValueError):
            return None
        if not (33.0 <= lat <= 39.0 and 124.0 <= lng <= 132.0):
            return None

        return self._build_poi_input(raw, lat, lng)

    async def _normalize_async(self, raw: dict) -> POIInput | None:
        if not self._is_in_district(raw):
            return None

        # Try direct coords first.
        try:
            lat = float(raw.get(COL_LAT))
            lng = float(raw.get(COL_LNG))
            if 33.0 <= lat <= 39.0 and 124.0 <= lng <= 132.0:
                return self._build_poi_input(raw, lat, lng)
        except (TypeError, ValueError):
            pass

        # Fall back to geocoding the address.
        addr = (raw.get(COL_ROAD_ADDR) or raw.get(COL_LOT_ADDR) or "").strip()
        if not addr:
            return None
        coords = await self._geocode(addr)
        if coords is None:
            return None
        return self._build_poi_input(raw, *coords)

    def _build_poi_input(self, raw: dict, lat: float, lng: float) -> POIInput:
        ext_id = _row_external_id(raw)
        attributes = {
            "accessibility": _accessibility_from_counts(raw),
            "gender": _gender_from_counts(raw),
            "opening_hours": _opening_hours(raw),
            "is_free": True,  # Public toilets are free unless flagged.
            "has_baby_changing": _truthy(raw.get(COL_BABY)),
            "address": (raw.get(COL_ROAD_ADDR) or raw.get(COL_LOT_ADDR) or None),
        }
        attributes = {k: v for k, v in attributes.items() if v is not None}

        return POIInput(
            external_id=ext_id,
            poi_type=POIType.toilet,
            lat=lat,
            lng=lng,
            name=(raw.get(COL_NAME) or "").strip() or None,
            attributes=attributes,
            last_verified_at=_parse_as_of(raw.get(COL_AS_OF)),
        )

    # --- run override (mirrors smoking_areas) ------------------------------

    async def run(self, session):  # type: ignore[override]
        report = ImportReport(source_id=self.source_id)
        run_started_at = datetime.now(timezone.utc)

        try:
            raw_rows = await self.fetch_raw()
        except Exception as e:  # noqa: BLE001
            report.errors.append(f"fetch_raw failed: {e}")
            logger.exception("fetch_raw failed for %s", self.source_id)
            return report

        skipped_no_coords = 0
        for raw in raw_rows:
            try:
                item = await self._normalize_async(raw)
            except Exception as e:  # noqa: BLE001
                report.errors.append(f"normalize failed: {e}")
                continue
            if item is None:
                # Track the most common reason — missing coords with no geocoder —
                # so the operator sees how much the run is leaving on the table.
                if self._is_in_district(raw) and not self._has_valid_coords(raw):
                    skipped_no_coords += 1
                continue
            try:
                outcome = await self._upsert_one(session, item, run_started_at)
                if outcome == "created":
                    report.created += 1
                elif outcome == "updated":
                    report.updated += 1
                else:
                    report.unchanged += 1
            except Exception as e:  # noqa: BLE001
                report.errors.append(f"upsert failed for {item.external_id}: {e}")

        if skipped_no_coords:
            msg = (
                f"skipped {skipped_no_coords} rows with blank coords and no geocoder; "
                f"pass `geocoder=` to recover them"
            )
            report.errors.append(msg)
            logger.warning(msg)

        try:
            report.removed = await self._soft_delete_stale(session, run_started_at)
        except Exception as e:  # noqa: BLE001
            report.errors.append(f"soft-delete failed: {e}")

        await session.commit()
        logger.info(str(report))
        return report

    @staticmethod
    def _has_valid_coords(raw: dict) -> bool:
        try:
            lat = float(raw.get(COL_LAT))
            lng = float(raw.get(COL_LNG))
        except (TypeError, ValueError):
            return False
        return 33.0 <= lat <= 39.0 and 124.0 <= lng <= 132.0
